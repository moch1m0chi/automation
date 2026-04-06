import os
from openpyxl import load_workbook
import re
from datetime import datetime

#ファイル名の月繰り上げ
def increment_month_in_filename(file_name):
    match = re.search(r"(\d+)月", file_name)
    
    if match:
        month = int(match.group(1))
        new_month = month + 1
        
        # 置き換え
        new_name = re.sub(r"\d+月", f"{new_month}月", file_name)
        return new_name
    else:
        return file_name  # 月が見つからなければそのまま
    
#元データを入れるフォルダの指定
input_folder = "data"
#処理したファイルの出力先フォルダを指定、なければoutputフォルダを作る
output_folder = "output"
os.makedirs(output_folder, exist_ok=True)

#正規表現(数値)/(数値)回目をコンパイル
pattern = re.compile(r"(\d+)/(\d+回目)")


for file_name in os.listdir(input_folder):
    if file_name.endswith((".xlsx", ".xlsm")):
        file_path = os.path.join(input_folder, file_name)
        print(f"処理中: {file_name}")

        wb = load_workbook(file_path)
        
        
        for ws in wb.worksheets:  # 全シート対象
            # 「報酬額確定合意書」シートがある場合のみ処理
            if "報酬額確定合意書" in wb.sheetnames:
                ws_agreement = wb["報酬額確定合意書"]
    
            for row in ws_agreement.iter_rows():
                for cell in row:
                    value = cell.value

                    # ■ 日付型の場合
                    if isinstance(value, datetime):
                        year = value.year
                        month = value.month + 1

                        if month > 12:
                            month = 1
                            year += 1

                        # 日付更新（末日ズレ防止）
                        try:
                            cell.value = value.replace(year=year, month=month)
                        except:
                            # 例：31日 → 翌月に存在しない場合は月末に調整
                            import calendar
                            last_day = calendar.monthrange(year, month)[1]
                            cell.value = value.replace(year=year, month=month, day=last_day)

                    # ■ 文字列（"2026/03/31"）の場合
                    elif isinstance(value, str):
                        try:
                            dt = datetime.strptime(value, "%Y/%m/%d")
                    
                            year = dt.year
                            month = dt.month + 1

                            if month > 12:
                                month = 1
                                year += 1

                            try:
                                new_dt = dt.replace(year=year, month=month)
                            except:
                                import calendar
                                last_day = calendar.monthrange(year, month)[1]
                                new_dt = dt.replace(year=year, month=month, day=last_day)

                            cell.value = new_dt.strftime("%Y/%m/%d")

                        except:
                            pass  # 日付じゃない文字列は無視

            #n/24回目をn+1/24回目に書き換える
            for row in ws.iter_rows():  #行を抜き出し
                for cell in row:    #抜き出した行のセルを走査
                    value = cell.value  #valueにセルの値を代入

                    if isinstance(value, str):  #セルの値value)はstr型(文字列)？
                        match = pattern.search(value)   #Matchオブジェクト
                        if match:   #matchの中身があるとTrueとして判定される。NoneだとFalse扱い
                            left = int(match.group(1)) + 1  #matchオブジェクトのmatch(1)、ここでは(/d+)に相当する部分
                            right = match.group(2)

                            new_value = f"{left}/{right}"   #f文字列
                            print(f"{ws.title} {cell.coordinate}: {value} → {new_value}")

                            cell.value = new_value

        new_file_name = increment_month_in_filename(file_name)  #ファイル名の月を繰り上げ

        output_path = os.path.join(output_folder, new_file_name)
        wb.save(output_path)

print("完了！")